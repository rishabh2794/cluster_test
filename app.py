import math
import tempfile
import numpy as np
import pandas as pd
import streamlit as st
import geopandas as gpd
from sklearn.cluster import DBSCAN
import folium
from folium.plugins import MarkerCluster
from shapely.geometry import Point
from openpyxl import load_workbook
from streamlit_folium import st_folium

# Optional: improve KML handling if available
try:
    import fiona
    HAS_FIONA = True
except Exception:
    HAS_FIONA = False

# -------------------------
# Constants & Helpers
# -------------------------
EARTH_RADIUS_M = 6_371_000.0
REQUIRED_COLS = {
    'ISSUE ID', 'CITY', 'ZONE', 'WARD', 'SUBCATEGORY', 'CREATED AT',
    'STATUS', 'LATITUDE', 'LONGITUDE', 'BEFORE PHOTO', 'AFTER PHOTO', 'ADDRESS'
}

st.set_page_config(layout="wide")
st.title("Unified Clustering + Batch Navigation — Hotspots + Route to Next N Tickets")
st.title("Clustering Application v10")

with st.sidebar:
    st.markdown("### Tips")
    st.markdown(
        "- CSV must include LATITUDE/LONGITUDE in decimal degrees.\n"
        "- Clustering radius is in **meters** (converted to **radians** for haversine).\n"
        "- If KML ward reading fails, convert to GeoJSON and try again.\n"
        "- Select your starting location by clicking on the map in Step 4A."
    )

# -------------------------
# Utility functions
# -------------------------
def normalize_subcategory(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().str.lower()

def haversine_m(lat1, lon1, lat2, lon2) -> float:
    phi1 = math.radians(float(lat1)); phi2 = math.radians(float(lat2))
    dphi = math.radians(float(lat2) - float(lat1))
    dlmb = math.radians(float(lon2) - float(lon1))
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlmb/2)**2
    return 2 * EARTH_RADIUS_M * math.asin(math.sqrt(a))

def google_maps_url(origin_lat, origin_lon, dest_lat, dest_lon, mode="driving", waypoints=None) -> str:
    base = "https://www.google.com/maps/dir/"
    origin = f"{origin_lat},{origin_lon}"
    destination = f"{dest_lat},{dest_lon}"
    waypoint_str = "/".join([f"{lat},{lon}" for lat, lon in waypoints]) if waypoints else ""
    full_path = "/".join(filter(None, [origin, waypoint_str, destination]))
    return f"{base}{full_path}"

def hyperlinkify_excel(excel_path: str, sheet_name: str = "Clustering Application Summary") -> None:
    try:
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            for col_idx in (11, 12):  # BEFORE (K), AFTER (L)
                link = ws.cell(row, col_idx).value
                if link and isinstance(link, str) and link.startswith(("http://", "https://")):
                    ws.cell(row, col_idx).hyperlink = link
                    ws.cell(row, col_idx).style = "Hyperlink"
        wb.save(excel_path)
    except Exception as e:
        st.warning(f"Excel hyperlinking skipped: {e}")

def load_wards_uploaded(file) -> gpd.GeoDataFrame | None:
    """Read GeoJSON/JSON/KML wards into EPSG:4326. Return None on failure and show an error."""
    try:
        suffix = file.name.split(".")[-1].lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix="." + suffix) as tmp:
            tmp.write(file.read())
            tmp_path = tmp.name

        if suffix in ("geojson", "json"):
            wards = gpd.read_file(tmp_path)
        elif suffix == "kml":
            if HAS_FIONA:
                wards = gpd.read_file(tmp_path, driver='KML')
            else:
                st.error("KML reading requires `fiona`. Convert to GeoJSON instead.")
                return None
        else:
            st.error("Unsupported ward file type. Please upload GeoJSON/JSON/KML.")
            return None

        if wards is None or wards.empty:
            st.error("Ward file loaded but contains no features.")
            return None

        if wards.crs is None:
            st.info("Ward file had no CRS. Assuming EPSG:4326 (WGS84).")
            wards.set_crs(epsg=4326, inplace=True)
        elif wards.crs.to_string() != "EPSG:4326":
            wards = wards.to_crs(epsg=4326)

        wards = wards[wards.geometry.notna()].copy()
        wards = wards[~wards.geometry.is_empty]
        return wards
    except Exception as e:
        st.error(f"Error reading ward file: {e}")
        return None

# -------------------------
# Session State
# -------------------------
if "visited_ticket_ids" not in st.session_state:
    st.session_state.visited_ticket_ids = set()
if "skipped_ticket_ids" not in st.session_state:
    st.session_state.skipped_ticket_ids = set()
if "current_target_id" not in st.session_state:
    st.session_state.current_target_id = None
if "batch_target_ids" not in st.session_state:
    st.session_state.batch_target_ids = set()
if 'map_center' not in st.session_state:
    st.session_state.map_center = [12.9716, 77.5946]  # Default to Bengaluru
if 'origin_coords' not in st.session_state:
    st.session_state.origin_coords = None

# -------------------------
# UI — Inputs
# -------------------------
st.subheader("Step 1: Upload Required Files")
csv_file = st.file_uploader("Upload CSV file with issues", type=["csv"])
ward_file = st.file_uploader("Upload WARD boundary file (GeoJSON/JSON/KML, optional)", type=["geojson", "json", "kml"])

subcategory_options = [
    "Pothole",
    "Sand piled on roadsides + Mud/slit on roadside",
    "Garbage Vulnerable Point",
    "Garbage vulnerable points",
    "Garbage Vulnerable points",
    "Open Burning",
    "Dumping of construction and demolition waste",
    "Other road damage",
    "Greening of Central Verges",
    "Unsurfaced Parking Lots"
]
st.subheader("Step 2: Select Issue Subcategory")
subcategory_option = st.selectbox("Choose issue subcategory to analyze:", subcategory_options)

# Parameters
st.subheader("Step 3: Set Clustering Parameters")
radius_m = st.number_input("Clustering Radius (meters)", min_value=1, max_value=1000, value=15)
min_samples = st.number_input("Minimum Issues per Cluster", min_value=1, max_value=100, value=2)

if radius_m < 10 or min_samples < 2:
    st.warning("⚠️ Low values may lead to too many tiny clusters. Proceed with caution.")

# -------------------------
# Core logic
# -------------------------
if csv_file:
    try:
        df = pd.read_csv(csv_file)
        missing = sorted(list(REQUIRED_COLS - set(df.columns)))
        if missing:
            st.error(f"Missing required columns: {missing}")
            st.stop()

        if not df.empty and 'LATITUDE' in df.columns and 'LONGITUDE' in df.columns:
            st.session_state.map_center = [df['LATITUDE'].mean(), df['LONGITUDE'].mean()]

        # Clean + filter by subcategory
        df['SUBCATEGORY_NORM'] = normalize_subcategory(df['SUBCATEGORY'])
        desired = subcategory_option.strip().lower()
        df = df[df['SUBCATEGORY_NORM'] == desired].copy()
        if df.empty:
            st.info("No rows found for the selected subcategory.")
            st.stop()

        df['LATITUDE'] = pd.to_numeric(df['LATITUDE'], errors='coerce')
        df['LONGITUDE'] = pd.to_numeric(df['LONGITUDE'], errors='coerce')
        df = df.dropna(subset=['LATITUDE', 'LONGITUDE'])
        if df.empty:
            st.info("All rows had invalid/missing coordinates after cleaning.")
            st.stop()

        for col in ['CREATED AT', 'STATUS', 'ADDRESS', 'BEFORE PHOTO', 'AFTER PHOTO',
                    'ISSUE ID', 'ZONE', 'WARD', 'CITY']:
            if col in df.columns:
                df[col] = df[col].astype(str)

        # --- DBSCAN ---
        coords_rad = np.radians(df[['LATITUDE', 'LONGITUDE']].to_numpy())
        eps_rad = float(radius_m) / EARTH_RADIUS_M
        db = DBSCAN(eps=eps_rad, min_samples=int(min_samples), metric='haversine', algorithm='ball_tree')
        labels = db.fit_predict(coords_rad)
        df['CLUSTER NUMBER'] = labels
        df['IS_CLUSTERED'] = df['CLUSTER NUMBER'] != -1

        # Optional ward join
        wards_gdf = load_wards_uploaded(ward_file) if ward_file else None
        gdf_all = gpd.GeoDataFrame(
            df.copy(),
            geometry=gpd.points_from_xy(df['LONGITUDE'].astype(float), df['LATITUDE'].astype(float)),
            crs="EPSG:4326"
        )
        if wards_gdf is not None and not wards_gdf.empty:
            try:
                gdf_all = gpd.sjoin(gdf_all, wards_gdf, how="left", predicate="within")
            except Exception as e:
                st.warning(f"Spatial join failed; proceeding without ward attribution. Error: {e}")

        # --- Summary dataframe ---
        clustered = gdf_all[gdf_all['IS_CLUSTERED']].copy()
        if not clustered.empty:
            sizes = clustered.groupby('CLUSTER NUMBER')['ISSUE ID'].count().rename("NUMBER OF ISSUES")
            summary_sheet = clustered.drop(columns=['SUBCATEGORY_NORM']).copy()
            summary_sheet = summary_sheet.merge(sizes, on='CLUSTER NUMBER', how='left')
            summary_sheet = summary_sheet[[
                'CLUSTER NUMBER', 'NUMBER OF ISSUES', 'ISSUE ID', 'ZONE', 'WARD',
                'SUBCATEGORY', 'CREATED AT', 'STATUS', 'LATITUDE', 'LONGITUDE',
                'BEFORE PHOTO', 'AFTER PHOTO', 'ADDRESS'
            ]].sort_values(['CLUSTER NUMBER', 'CREATED AT'])
        else:
            summary_sheet = pd.DataFrame(columns=[
                'CLUSTER NUMBER', 'NUMBER OF ISSUES', 'ISSUE ID', 'ZONE', 'WARD',
                'SUBCATEGORY', 'CREATED AT', 'STATUS', 'LATITUDE', 'LONGITUDE',
                'BEFORE PHOTO', 'AFTER PHOTO', 'ADDRESS'
            ])

        # Save Excel
        excel_filename = "Clustering_Application_Summary.xlsx"
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            summary_sheet.to_excel(writer, index=False, sheet_name='Clustering Application Summary')
        hyperlinkify_excel(excel_filename)

        # -------------------------
        # Step 4A — Batch Navigation
        # -------------------------
        st.subheader("Step 4A: Navigate to Nearest Tickets (Batch)")
        unique_statuses = sorted(gdf_all['STATUS'].dropna().astype(str).unique().tolist())
        default_statuses = [s for s in unique_statuses if s.lower() in ("open", "pending", "in progress")]
        include_statuses = st.multiselect("Eligible ticket statuses", options=unique_statuses, default=default_statuses or unique_statuses)
        wards_in_data = sorted(gdf_all['WARD'].dropna().astype(str).unique().tolist())
        ward_filter = st.multiselect("Limit to ward(s) (optional)", options=wards_in_data, default=[])
        travel_mode = st.selectbox("Travel mode", ["driving", "walking", "two_wheeler"], index=0)
        batch_size = st.slider("Batch size (next N tickets)", min_value=1, max_value=10, value=10)

        # Map for origin
        st.markdown("### Your Location")
        m = folium.Map(location=st.session_state.map_center, zoom_start=12)
        if st.session_state.origin_coords:
            folium.Marker(
                location=[st.session_state.origin_coords['lat'], st.session_state.origin_coords['lng']],
                popup="Your Starting Location",
                icon=folium.Icon(color="green"),
            ).add_to(m)
        map_data = st_folium(m, width=725, height=400)
        if map_data and map_data.get("last_clicked"):
            st.session_state.origin_coords = map_data["last_clicked"]

        origin_lat = origin_lon = None
        if st.session_state.origin_coords:
            origin_lat = st.session_state.origin_coords['lat']
            origin_lon = st.session_state.origin_coords['lng']
            st.success(f"✅ Starting location set: {origin_lat:.6f}, {origin_lon:.6f}")

        # Build pool
        pool = gdf_all.copy()
        if include_statuses:
            pool = pool[pool['STATUS'].astype(str).isin(include_statuses)]
        if ward_filter:
            pool = pool[pool['WARD'].astype(str).isin([str(w) for w in ward_filter])]
        if st.session_state.visited_ticket_ids:
            pool = pool[~pool['ISSUE ID'].astype(str).isin(st.session_state.visited_ticket_ids)]
        if st.session_state.skipped_ticket_ids:
            pool = pool[~pool['ISSUE ID'].astype(str).isin(st.session_state.skipped_ticket_ids)]
        st.write(f"Eligible tickets remaining: **{len(pool)}**")

        # Nearest-neighbor sequence
        sequence_rows, leg_dists = [], []
        if origin_lat is not None and origin_lon is not None and not pool.empty:
            cur_lat, cur_lon = origin_lat, origin_lon
            pool2 = pool.copy()
            for _ in range(min(batch_size, len(pool2))):
                pool2['__dist_m'] = pool2.apply(lambda r: haversine_m(cur_lat, cur_lon, r['LATITUDE'], r['LONGITUDE']), axis=1)
                nxt = pool2.sort_values('__dist_m', ascending=True).iloc[0]
                sequence_rows.append(nxt)
                cur_lat, cur_lon = float(nxt['LATITUDE']), float(nxt['LONGITUDE'])
                pool2 = pool2[pool2['ISSUE ID'] != nxt['ISSUE ID']]

        if sequence_rows:
            waypoints = [(float(r['LATITUDE']), float(r['LONGITUDE'])) for r in sequence_rows]
            last_stop = waypoints[-1]; mid_waypoints = waypoints[:-1]
            nav_url = google_maps_url(origin_lat, origin_lon, last_stop[0], last_stop[1], mode=travel_mode, waypoints=mid_waypoints)
            total_m, prev = 0.0, (origin_lat, origin_lon)
            for (lat, lon) in waypoints:
                d = haversine_m(prev[0], prev[1], lat, lon); leg_dists.append(int(d)); total_m += d; prev = (lat, lon)
            eta_min = total_m / (30 * 1000 / 60)

            st.success(
                f"Batch route ready: **{len(sequence_rows)}** tickets | "
                f"Total distance ≈ **{int(total_m)} m** | ETA ≈ **{eta_min:.1f} min**"
            )
            st.markdown(f"[🧭 Open continuous navigation in Google Maps]({nav_url})")

            # Ordered list
            list_df = pd.DataFrame({
                "#": list(range(1, len(sequence_rows)+1)),
                "ISSUE ID": [str(r['ISSUE ID']) for r in sequence_rows],
                "WARD": [str(r.get('WARD', '')) for r in sequence_rows],
                "STATUS": [str(r.get('STATUS', '')) for r in sequence_rows],
                "Leg dist (m)": leg_dists
            })
            st.dataframe(list_df, use_container_width=True)

            st.session_state.current_target_id = str(sequence_rows[0]['ISSUE ID'])
            st.session_state.batch_target_ids = {str(r['ISSUE ID']) for r in sequence_rows}

            c1, c2 = st.columns(2)
            with c1:
                if st.button("✅ Mark first as Done (visited)"):
                    st.session_state.visited_ticket_ids.add(str(sequence_rows[0]['ISSUE ID']))
                    st.rerun()
            with c2:
                if st.button("⏭️ Skip first ticket"):
                    st.session_state.skipped_ticket_ids.add(str(sequence_rows[0]['ISSUE ID']))
                    st.rerun()
        else:
            st.info("Set your starting location on the map to compute a batch route.")

        # -------------------------
        # Step 5 — Map Display
        # -------------------------
        st.subheader("Step 5: Map Display Options")
        center_on_first = st.checkbox("Center map on first stop (if any)", value=True)
        map_type = st.radio("Select map type:", ["Show all markers (Type 1)", "Use Dynamic Clustering (Type 2)"], index=0)
        if center_on_first and sequence_rows:
            display_map_center = [float(sequence_rows[0]['LATITUDE']), float(sequence_rows[0]['LONGITUDE'])]; zoom_level = 16
        else:
            display_map_center, zoom_level = st.session_state.map_center, 13
        display_map = folium.Map(location=display_map_center, zoom_start=zoom_level)

        if wards_gdf is not None and not wards_gdf.empty:
            try: folium.GeoJson(wards_gdf, name="Wards").add_to(display_map)
            except Exception: pass

        if origin_lat is not None:
            folium.Marker(location=[origin_lat, origin_lon], popup="Your Starting Location",
                          icon=folium.Icon(color="green", icon="star")).add_to(display_map)

        target_id, batch_ids = st.session_state.get('current_target_id'), st.session_state.get('batch_target_ids', set())
        ticket_markers = folium.FeatureGroup(name="Ticket Markers")
        if map_type == "Show all markers (Type 1)":
            for _, row in gdf_all.iterrows():
                rid = str(row['ISSUE ID']); is_first = (rid == str(target_id)) if target_id else False; in_
